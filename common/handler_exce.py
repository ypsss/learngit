import openpyxl
import os
from common.heandler_path import DATA_DTR


class HeadleExcel:

    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        """读取excel"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        res = list(sh.rows)
        '''获取第一行表头'''
        title = [i.value for i in res[0]]
        cases = []
        """遍历第一行之外的其他行"""
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        """获取返回的数据"""
        return cases 

    def write_data(self, row, column, value):
        """写入excel"""
        workbook = openpyxl.load_workbook(self.filename)
        sh = workbook[self.sheetname]
        sh.cell(row=row, column=column, value=value)
        workbook.save(self.filename)


if __name__ == '__main__':
    cc = HeadleExcel(r"C:\Users\ZLD\PycharmProjects\Learn\datas\testcase.xlsx","loging")
    print(cc.read_data())
